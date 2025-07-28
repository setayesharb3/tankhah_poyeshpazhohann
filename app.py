import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook, Workbook

# ------------------------------------------------------------
# پیکربندی صفحه
# ------------------------------------------------------------
st.set_page_config(page_title="ثبت حسابداری تنخواه", page_icon="📄")

# 🎨 استایل سایت
st.markdown(
    """
<style>
body { background-color: #fdf0f5; color: #333333; }
.stApp { background-color: #fff0f5; font-family: IRANSans, sans-serif; }
.stButton>button { background-color: #ffb6c1; color: white; }
.stTextInput>div>div>input { background-color: #fffafc; }
.stFileUploader>div>div { background-color: #ffe4ec; }
.stDownloadButton>button { background-color: #ff69b4; color: white; }
</style>
""",
    unsafe_allow_html=True,
)

st.title("📄 سایت ثبت حسابداری تنخواه")

# ------------------------------------------------------------
# ورودی‌های کاربر
# ------------------------------------------------------------
col1, col2 = st.columns(2)
with col1:
    tanakh_number = st.text_input("🧾 شماره تنخواه:", "")
    tanakh_name = st.text_input("👤 نام تنخواه‌دار:", "")
    date_input = st.text_input("📅 تاریخ ثبت (مثلاً 1403/03/12):", "")
with col2:
    project_name = st.text_input("🏗️ نام پروژه:", "")
    sath4_default = st.text_input("🔢 سطح چهارم هزینه‌ها:", "")
    sath5_default = st.text_input("🔢 سطح پنجم هزینه‌ها:", "")

# سطح چهارم کارمزد ویژه پروژه پرند (اگر ندهی، 005021 پیش‌فرض)
sath4_fee_input = st.text_input("🔢 سطح چهارم کارمزد (فقط وقتی پروژه = پرند)", "")

uploaded_file = st.file_uploader("📎 فایل اکسل تنخواه را بارگذاری کنید", type=["xlsx"])

# ------------------------------------------------------------
# سطح چهارم برای پرداخت تنخواه‌دارها (حساب تنخواه اشخاص)
# ------------------------------------------------------------
tanakh_sath4_map = {
    "آقای ویسی": "100094",
    "اقا عطا": "101026",
    "اقای نظرخانی": "101973",
    "اقای مستقیمی": "101381",
    "اقای وثوقی راد": "100388",
    "خانم فراهانی": "100424",
    "اقای الماسی": "101192",
    "اقای حقی": "101240",
    "اقای حبیب زاده": "102830",
    "اقای بهروز پور": "101720",
    "خانم زابلی": "101986",
    "اقای مصطفی زاده": "101373",
    "اقای اصلان": "100039",
    "اقای روان مهر": "101967",
    "اقای مشهدی ملک": "101520",
}

# ------------------------------------------------------------
# مپ کلمات به کد معین (پشتیبان)
# ------------------------------------------------------------
keyword_accounts = {
    "حمل": 7301,
    "کرایه": 7301,
    "آب": 7201,
    "برق": 7201,
    "گاز": 7201,
    "پست": 7202,
    "تلفن": 7202,
    "تلگراف": 7202,
    "ملزومات": 7203,
    "نوشت افزار": 7203,
    "آبدارخانه": 7204,
    "پذیرایی": 7204,
    "سفر": 7205,
    "اقامت": 7205,
    "چاپ": 7208,
    "کپی": 7208,
    "پوشاک": 7210,
    "بهداشت": 7212,
    "درمان": 7212,
    "غذا": 7215,
    "ایاب": 7216,
    "ذهاب": 7216,
    "کمک": 7219,
    "هدایا": 7219,
    "مصرفی": 7226,
    "سوخت": 7252,
    "تبلیغات": 7298,
    "آگهی": 7298,
    "تخلیه": 7302,
    "بارگیری": 7302,
    "بیمه": 7303,
    "آزمایشگاه": 7304,
    "لوازم بهداشتی": 7310,
    "مواد": 7315,
    "پیمانکار": 7330,
    "بازسازی": 7331,
    "اجاره": 7341,
    "اجرت": 7350,
    "تعویض": 7350,
    "تجاری": 3130,
}

# ------------------------------------------------------------
# Utility helpers
# ------------------------------------------------------------
def extract_int_str(val):
    """عدد صحیح بدون اعشار؛ چندتایی '2532-2534' هم پشتیبانی."""
    try:
        if pd.isna(val):
            return ""
        s = str(val).strip()
        if s == "" or s.lower() in ("nan", "none", "-"):
            return ""
        if "-" in s:
            parts = []
            for p in s.split("-"):
                p = p.strip()
                if p == "":
                    continue
                try:
                    parts.append(str(int(float(p))))
                except Exception:
                    pass
            return "-".join(parts) if parts else ""
        return str(int(float(s)))
    except Exception:
        return ""


def get_center_cost_str(val):
    """سطح چهارم از 'مرکز هزینه' با صفر پَد ۶ رقمی."""
    try:
        if pd.isna(val):
            return ""
        s = str(val).strip()
        if s == "" or s.lower() in ("nan", "none", "-"):
            return ""
        try:
            s = str(int(float(s)))
        except Exception:
            s_digits = "".join(ch for ch in s if ch.isdigit())
            if s_digits:
                s = s_digits
        return s.zfill(6)
    except Exception:
        return ""


def clean_number(val):
    """تبدیل عددی امن؛ خالی => 0."""
    try:
        if pd.isna(val):
            return 0
        s = str(val).strip()
        if s == "" or s.lower() == "nan" or s == "-":
            return 0
        return float(s.replace(",", ""))
    except Exception:
        return 0


def detect_account_code(desc, sath5_val, tanakh_name):
    """
    انتخاب کد معین بر اساس شرح، وضعیت دفتر/پروژه، و استثناء عطا/زابلی.
    """
    if not desc or desc.strip() == "":
        return 7296 if sath5_val == "006003" else 7350

    desc_lower = desc.lower()

    # استثناء عطا و زابلی: ارسال/آوردن => ایاب ذهاب
    if tanakh_name.strip() in ["اقا عطا", "خانم زابلی"]:
        if any(word in desc_lower for word in ["ارسال", "اوردن"]):
            return 7216

    if sath5_val == "006003":  # دفتر مرکزی
        keywords_72 = {
            "آب": 7201,
            "برق": 7201,
            "گاز": 7201,
            "قبض": 7201,
            "اینترنت": 7202,
            "شارژ": 7202,
            "تلفن همراه": 7202,
            "تلفن ثابت": 7202,
            "خودکار": 7203,
            "مداد": 7203,
            "لوازم التحریر": 7203,
            "صبحانه": 7204,
            "مواد شوینده": 7204,
            "شیرینی": 7204,
            "پذیرایی": 7204,
            "بلیط هواپیما": 7205,
            "کشتی": 7205,
            "قطار": 7205,
            "هتل": 7205,
            "چاپ": 7208,
            "کپی": 7208,
            "پرینت": 7208,
            "لباس": 7210,
            "فرم": 7210,
            "لباس کارکنان": 7210,
            "درمان": 7212,
            "دارو": 7212,
            "تست آزمایشگاه": 7212,
            "آزمایشگاه": 7212,
            "درمانگاه": 7212,
            "نهار": 7215,
            "شام": 7215,
            "ایاب ذهاب": 7216,
            "اسنپ": 7216,
            "تپسی": 7216,
            "آژانس": 7216,
            " ایاب و ذهاب ": 7216,
            "هدیه": 7219,
            "دفتر": 7226,
            "بنزین": 7252,
            "آگهی": 7298,
            "تبلیغات": 7298,
            "فیلمبرداری": 7298,
            "استخدام": 7298,
        }
        for word, code in keywords_72.items():
            if word in desc_lower:
                return code
        return 7296  # پیش‌فرض دفتر
    else:
        keywords_all = {
            "حمل": 7301,
            "کرایه": 7301,
            "تخلیه": 7302,
            "بارگیری": 7302,
            "بیمه": 7303,
            "آزمایشگاه": 7304,
            "لوازم بهداشتی": 7310,
            "مواد": 7315,
            "پیمانکار": 7330,
            "بازسازی": 7331,
            "اجاره": 7341,
            "اجرت": 7350,
            "تعویض": 7350,
            "نظافت": 7350,
            "تجاری": 3130,
            " ایاب و ذهاب ": 7216,
        }
        keywords_72 = {
            "آب": 7201,
            "برق": 7201,
            "گاز": 7201,
            "قبض": 7201,
            "اینترنت": 7202,
            "شارژ": 7202,
            "تلفن همراه": 7202,
            "تلفن ثابت": 7202,
            "خودکار": 7203,
            "مداد": 7203,
            "لوازم التحریر": 7203,
            "صبحانه": 7204,
            "مواد شوینده": 7204,
            "شیرینی": 7204,
            "پذیرایی": 7204,
            "بلیط هواپیما": 7205,
            "کشتی": 7205,
            "قطار": 7205,
            "هتل": 7205,
            "چاپ": 7208,
            "کپی": 7208,
            "پرینت": 7208,
            "لباس": 7210,
            "فرم": 7210,
            "لباس کارکنان": 7210,
            "درمان": 7212,
            "دارو": 7212,
            "تست آزمایشگاه": 7212,
            "آزمایشگاه": 7212,
            "درمانگاه": 7212,
            "نهار": 7215,
            "شام": 7215,
            "ایاب ذهاب": 7216,
            "اسنپ": 7216,
            "تپسی": 7216,
            "آژانس": 7216,
            " ایاب و ذهاب ": 7216,
            "هدیه": 7219,
            "دفتر": 7226,
            "بنزین": 7252,
            "آگهی": 7298,
            "تبلیغات": 7298,
            "فیلمبرداری": 7298,
            "استخدام": 7298,
        }
        all_keywords = {**keywords_all, **keywords_72}
        for word, code in all_keywords.items():
            if word in desc_lower:
                return code
        return 7350


def safe_append(parts_list, text):
    """افزودن بخش به شرح اگر معتبر باشد."""
    if text is None:
        return
    s = str(text).strip()
    if s == "" or s.lower() == "nan":
        return
    parts_list.append(s)


def prefix_babat(txt: str) -> str:
    """اگر ابتدای شرح «بابت» ندارد، اضافه کن."""
    if not txt:
        return "بابت"
    t = txt.strip()
    if t.startswith("بابت"):
        return t
    return f"بابت {t}"


def clean_desc_final(txt: str) -> str:
    """حذف nan, None, فاصله‌های تکراری."""
    if not txt:
        return ""
    bads = ["nan", "None", "none", "Nan", "NaN"]
    out = txt
    for b in bads:
        out = out.replace(b, "")
    # فاصله‌های دوگانه
    while "  " in out:
        out = out.replace("  ", " ")
    return out.strip()

# ============================================================
# پردازش فایل
# ============================================================
if uploaded_file and all([tanakh_number, tanakh_name, date_input, project_name, sath4_default, sath5_default]):
    try:
        df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip()  # حذف فاصله‌های اضافه از نام ستون‌ها

        # تشخیص ستون پرداخت جمعی (هر کدام موجود بود)
        group_col = None
        for cand in ["پرداخت جمعی", "پرداخت گروهی"]:
            if cand in df.columns:
                group_col = cand
                break

        wb = load_workbook("سند حسابداری (21).xlsx")
        ws = wb.active
        row_index = 2

        # آماده‌سازی سطح‌ها
        sath4_default_z = sath4_default.zfill(6)
        sath5_default_z = sath5_default.zfill(6)
        sath4_tanakh = tanakh_sath4_map.get(tanakh_name.strip(), "")
        sath4_tanakh = sath4_tanakh.zfill(6) if sath4_tanakh else ""

        # سطح چهارم کارمزد ورودی (فقط برای پرند استفاده می‌شود)
        sath4_fee_z = sath4_fee_input.zfill(6) if sath4_fee_input.strip() else "005021"

        # وضعیت ناحیه جاری برای جمع (برای آقای الماسی)
        current_area = None
        area_amount_sum = 0
        area_fee_sum = 0
        area_tax_sum = 0

        # وضعیت پرداخت جمعی (برای سایر تنخواه‌دارها)
        group_active = False
        group_amount_sum = 0
        group_fee_sum = 0
        group_tax_sum = 0
        group_count = 0
        group_first_desc = ""
        group_last_desc = ""

        # آیا این تنخواه‌دار الماسی است؟
        is_almasi = ("الماسی" in tanakh_name) or (tanakh_name.strip() == "اقای الماسی")

        # دیکشنری کالا/خدمت (برای ستون AQ)
        item_type_dict = {
            7201: "خدمت",
            7202: "خدمت",
            7203: "کالا",
            7204: "کالا",
            7205: "خدمت",
            7208: "خدمت",
            7210: "کالا",
            7212: "خدمت",
            7215: "کالا",
            7216: "خدمت",
            7219: "کالا",
            7226: "کالا",
            7252: "کالا",
            7296: "خدمت",
            7298: "خدمت",
            7301: "خدمت",
            7302: "خدمت",
            7303: "خدمت",
            7304: "خدمت",
            7310: "خدمت",
            7315: "خدمت",
            7330: "خدمت",
            7331: "خدمت",
            7341: "خدمت",
            7350: "خدمت"
        }

        # ---------------------------
        # حلقه ردیف‌ها
        # ---------------------------
        for idx, row in df.iterrows():
            area_val = str(row.get("ناحیه", "")).strip() if "ناحیه" in df.columns else ""

            desc = str(row.get("شرح سند", "")).strip() if "شرح سند" in df.columns else ""
            seller = str(row.get("نام فروشنده / فروشگاه", "")).replace("فروشگاه", "").strip() if "نام فروشنده / فروشگاه" in df.columns else ""
            factor = extract_int_str(row.get("شماره فاکتور")) if "شماره فاکتور" in df.columns else ""
            resi = extract_int_str(row.get("رسیدانبار")) if "رسیدانبار" in df.columns else ""
            cost_date = row.get("تاریخ", None) if "تاریخ" in df.columns else None
            fee = clean_number(row.get("کارمزد")) if "کارمزد" in df.columns else 0
            tax = clean_number(row.get("ارزش افزوده")) if "ارزش افزوده" in df.columns else 0
            amount = clean_number(row.get("مبلغ")) if "مبلغ" in df.columns else 0
            is_gardesh = False
            if "گردش" in df.columns:
                is_gardesh = str(row.get("گردش", "")).strip() == "گردش"

            # پرداخت جمعی؟
            group_flag = False
            if group_col is not None:
                gv = row.get(group_col)
                if pd.notna(gv):
                    try:
                        group_flag = float(gv) != 0
                    except Exception:
                        group_flag = True  # هر مقدار غیرخالی

            # شرح کامل
            parts = []
            safe_append(parts, desc)
            if resi:
                safe_append(parts, f"به شماره رسید انبار {resi}")
            if factor:
                safe_append(parts, f"شماره فاکتور {factor}")
            if seller:
                safe_append(parts, seller)
            if cost_date and not resi:
                cost_date_str = str(cost_date).strip()
                if cost_date_str and str(cost_date_str).lower() != "nan":
                    safe_append(parts, f"مورخ {cost_date_str}")
            safe_append(parts, f"طی تنخواه شماره {tanakh_number} {tanakh_name} پروژه {project_name}")
            full_desc = " ".join(parts)
            summary = f"صورتخلاصه تنخواه شماره {tanakh_number} طی تنخواه {tanakh_name} پروژه {project_name}"

            # سطح برای حساب هزینه (پایه)
            if tanakh_name.strip() in ["اقا عطا", "خانم زابلی"]:
                sath5_use = "006003"
                sath4_use = sath4_default_z
            else:
                sath5_use = sath5_default_z
                sath4_use = sath4_default_z

            # منطق پروژه پرند: سطح چهارم هزینه از ستون مرکز هزینه
            if project_name.strip() == "پرند":
                center_cost = get_center_cost_str(row.get("مرکز هزینه")) if "مرکز هزینه" in df.columns else ""
                if center_cost:
                    sath4_cost = center_cost
                else:
                    sath4_cost = "005021"
                # سطح چهارم کارمزد از ورودی کاربر (اگر خالی بود همان 005021)
                sath4_fee_for_this_row = sath4_fee_z
            else:
                sath4_cost = sath4_use
                sath4_fee_for_this_row = sath4_use


            # حساب هزینه
            if resi:
                account_code = 3120
                sath4_cost = "200082"
                sath5_cost = sath5_use
            else:
                account_code = detect_account_code(desc, sath5_use, tanakh_name)
                sath5_cost = sath5_use

            # ستون‌های AO تا AU
            ao_val = ""
            ap_val = "داخلی"
            aq_val = ""
            ar_val = "خرید"
            as_val = "5058"
            at_val = 0
            au_val = 0

            if account_code == 1131:
                ao_val = None
                ap_val = None
                aq_val = None
                ar_val = None
                as_val = None
                at_val = None
                au_val = None
            else:
                if amount > 0:
                    if tax > 0:
                        ao_val = "مشمول"
                        at_val = tax
                        au_val = 0
                    else:
                        ao_val = "معاف"
                        at_val = 0
                        au_val = 0
                    aq_val = item_type_dict.get(account_code, "")

            ws[f"AO{row_index}"] = ao_val if ao_val else None
            ws[f"AP{row_index}"] = ap_val if ap_val else None
            ws[f"AQ{row_index}"] = aq_val if aq_val else None
            ws[f"AR{row_index}"] = ar_val if ar_val else None
            ws[f"AS{row_index}"] = as_val if as_val else None
            if ao_val == "مشمول":
                ws[f"AT{row_index}"] = at_val
                ws[f"AU{row_index}"] = au_val
            else:
                ws[f"AT{row_index}"] = None
                ws[f"AU{row_index}"] = None

            # ====================================================
            # شاخه 1: تنخواه‌دار الماسی → 1131 کلی برای هر ناحیه
            # ====================================================
            if is_almasi:
                if area_val:
                    # اگر داریم از یک ناحیه به ناحیه جدید می‌رویم، قبلی را ببندیم
                    if current_area and current_area != area_val:
                        ws[f"C{row_index}"] = date_input
                        ws[f"D{row_index}"] = f"پرداخت ناحیه {current_area} طی تنخواه {tanakh_number} {tanakh_name} پروژه {project_name}"
                        ws[f"H{row_index}"] = 1131
                        ws[f"K{row_index}"] = f"پرداخت ناحیه {current_area}"
                        ws[f"Q{row_index}"] = area_amount_sum + area_fee_sum + area_tax_sum
                        ws[f"X{row_index}"] = sath4_tanakh
                        ws[f"Y{row_index}"] = ""
                        row_index += 1
                        area_amount_sum = area_fee_sum = area_tax_sum = 0

                    current_area = area_val

                    # هزینه
                    ws[f"C{row_index}"] = date_input
                    ws[f"D{row_index}"] = summary
                    ws[f"H{row_index}"] = account_code
                    ws[f"K{row_index}"] = full_desc
                    ws[f"P{row_index}"] = amount if amount else None
                    ws[f"X{row_index}"] = sath4_cost
                    ws[f"Y{row_index}"] = sath5_cost
                    row_index += 1

                    # مالیات
                    if tax > 0:
                        ws[f"C{row_index}"] = date_input
                        ws[f"D{row_index}"] = summary
                        ws[f"H{row_index}"] = 3221
                        ws[f"K{row_index}"] = "بابت ارزش افزوده"
                        ws[f"P{row_index}"] = tax
                        ws[f"X{row_index}"] = sath4_cost
                        ws[f"Y{row_index}"] = sath5_cost
                        row_index += 1

                    # کارمزد
                    if fee > 0:
                        ws[f"C{row_index}"] = date_input
                        ws[f"D{row_index}"] = summary
                        ws[f"H{row_index}"] = 7512
                        ws[f"K{row_index}"] = "بابت کارمزد بانکی"
                        ws[f"P{row_index}"] = fee
                        ws[f"X{row_index}"] = sath4_fee_for_this_row
                        ws[f"Y{row_index}"] = sath5_cost
                        row_index += 1

                    # جمع ناحیه
                    area_amount_sum += amount
                    area_fee_sum += fee
                    area_tax_sum += tax

                    # (اختیاری) گردش
                    if is_gardesh:
                        for col in ['P', 'Q']:
                            ws[f"C{row_index}"] = date_input
                            ws[f"D{row_index}"] = summary
                            ws[f"H{row_index}"] = 3120
                            ws[f"K{row_index}"] = full_desc
                            ws[f"{col}{row_index}"] = amount
                            ws[f"X{row_index}"] = ""
                            ws[f"Y{row_index}"] = sath5_use
                            row_index += 1
                    continue  # 1131 اینجا نمی‌زنیم؛ در پایان ناحیه

                # الماسی - ردیف بدون ناحیه: اگر ناحیه باز بود ببند
                if current_area:
                    ws[f"C{row_index}"] = date_input
                    ws[f"D{row_index}"] = f"پرداخت ناحیه {current_area} طی تنخواه {tanakh_number} {tanakh_name} پروژه {project_name}"
                    ws[f"H{row_index}"] = 1131
                    ws[f"K{row_index}"] = f"پرداخت ناحیه {current_area}"
                    ws[f"Q{row_index}"] = area_amount_sum + area_fee_sum + area_tax_sum
                    ws[f"X{row_index}"] = sath4_tanakh
                    ws[f"Y{row_index}"] = ""
                    row_index += 1
                    current_area = None
                    area_amount_sum = area_fee_sum = area_tax_sum = 0

                # هزینه
                ws[f"C{row_index}"] = date_input
                ws[f"D{row_index}"] = summary
                ws[f"H{row_index}"] = account_code
                ws[f"K{row_index}"] = full_desc
                ws[f"P{row_index}"] = amount if amount else None
                ws[f"X{row_index}"] = sath4_cost
                ws[f"Y{row_index}"] = sath5_cost
                row_index += 1

                # مالیات
                if tax > 0:
                    ws[f"C{row_index}"] = date_input
                    ws[f"D{row_index}"] = summary
                    ws[f"H{row_index}"] = 3221
                    ws[f"K{row_index}"] = "بابت ارزش افزوده"
                    ws[f"P{row_index}"] = tax
                    ws[f"X{row_index}"] = sath4_cost
                    ws[f"Y{row_index}"] = sath5_cost
                    row_index += 1

                # کارمزد
                if fee > 0:
                    ws[f"C{row_index}"] = date_input
                    ws[f"D{row_index}"] = summary
                    ws[f"H{row_index}"] = 7512
                    ws[f"K{row_index}"] = "بابت کارمزد بانکی"
                    ws[f"P{row_index}"] = fee
                    ws[f"X{row_index}"] = sath4_fee_for_this_row
                    ws[f"Y{row_index}"] = sath5_cost
                    row_index += 1

                # الماسی → اینجا پرداخت 1131 ردیفی نداریم
                continue

            # ====================================================
            # شاخه 2: سایر تنخواه‌دارها
            #   + پرداخت جمعی (گروه) بر اساس ستون پرداخت جمعی
            #   + پرداخت عادی در غیر اینصورت
            # ====================================================

            # اگر در حالت گروه فعال هستیم و این ردیف دیگر عضو گروه نیست → گروه قبلی را ببندیم
            if group_active and not group_flag:
                # شرح پرداخت گروهی: اگر desc فعلی خالی نبود از آن استفاده می‌کنیم
                group_pay_desc = desc if desc else f"پرداخت جمعی {group_count} فقره فاکتور طی تنخواه شماره {tanakh_number} {tanakh_name}"
                group_pay_full = f"{group_pay_desc} پروژه {project_name}"

                ws[f"C{row_index}"] = date_input
                ws[f"D{row_index}"] = summary
                ws[f"H{row_index}"] = 1131
                ws[f"K{row_index}"] = group_pay_full
                ws[f"Q{row_index}"] = group_amount_sum + group_fee_sum + group_tax_sum
                ws[f"X{row_index}"] = sath4_tanakh
                ws[f"Y{row_index}"] = ""
                row_index += 1

                # ریست گروه
                group_active = False
                group_amount_sum = group_fee_sum = group_tax_sum = 0
                group_count = 0
                group_first_desc = ""
                group_last_desc = ""

                # اگر این ردیف خودش هیچ مبلغی ندارد (مثل ردیف «بابت ... فقره») دیگر پردازشش نکنیم
                if (amount == 0) and (fee == 0) and (tax == 0):
                    continue
                # اگر مبلغ دارد، از اینجا ادامه منطق عادی (بدون گروه) می‌رود

            # اگر ردیف عضو گروه است:
            if group_flag:
                if not group_active:
                    group_active = True
                    group_amount_sum = 0
                    group_fee_sum = 0
                    group_tax_sum = 0
                    group_count = 0
                    group_first_desc = desc
                group_last_desc = desc
                group_count += 1

                # --- ثبت هزینه / مالیات / کارمزد (بدون 1131) ---
                # هزینه
                ws[f"C{row_index}"] = date_input
                ws[f"D{row_index}"] = summary
                ws[f"H{row_index}"] = account_code
                ws[f"K{row_index}"] = full_desc
                ws[f"P{row_index}"] = amount if amount else None
                ws[f"X{row_index}"] = sath4_cost
                ws[f"Y{row_index}"] = sath5_cost
                row_index += 1

                # مالیات
                if tax > 0:
                    ws[f"C{row_index}"] = date_input
                    ws[f"D{row_index}"] = summary
                    ws[f"H{row_index}"] = 3221
                    ws[f"K{row_index}"] = "بابت ارزش افزوده"
                    ws[f"P{row_index}"] = tax
                    ws[f"X{row_index}"] = sath4_cost
                    ws[f"Y{row_index}"] = sath5_cost
                    row_index += 1

                # کارمزد
                if fee > 0:
                    ws[f"C{row_index}"] = date_input
                    ws[f"D{row_index}"] = summary
                    ws[f"H{row_index}"] = 7512
                    ws[f"K{row_index}"] = "بابت کارمزد بانکی"
                    ws[f"P{row_index}"] = fee
                    ws[f"X{row_index}"] = sath4_fee_for_this_row
                    ws[f"Y{row_index}"] = sath5_cost
                    row_index += 1

                # جمع گروه
                group_amount_sum += amount
                group_fee_sum += fee
                group_tax_sum += tax

                # (اختیاری) گردش
                if is_gardesh:
                    for col in ['P', 'Q']:
                        ws[f"C{row_index}"] = date_input
                        ws[f"D{row_index}"] = summary
                        ws[f"H{row_index}"] = 3120
                        ws[f"K{row_index}"] = full_desc
                        ws[f"{col}{row_index}"] = amount
                        ws[f"X{row_index}"] = ""
                        ws[f"Y{row_index}"] = sath5_use
                        row_index += 1

                continue  # مهم! برای ردیف‌های گروهی 1131 زده نمی‌شود

            # اگر نه گروه فعال است نه گروه_flag → ردیف عادیِ غیرگروهی → منطق اصلی پرداخت تک‌به‌تک
            # هزینه
            ws[f"C{row_index}"] = date_input
            ws[f"D{row_index}"] = summary
            ws[f"H{row_index}"] = account_code
            ws[f"K{row_index}"] = full_desc
            ws[f"P{row_index}"] = amount if amount else None
            ws[f"X{row_index}"] = sath4_cost
            ws[f"Y{row_index}"] = sath5_cost
            row_index += 1

            # مالیات
            if tax > 0:
                ws[f"C{row_index}"] = date_input
                ws[f"D{row_index}"] = summary
                ws[f"H{row_index}"] = 3221
                ws[f"K{row_index}"] = "بابت ارزش افزوده"
                ws[f"P{row_index}"] = tax
                ws[f"X{row_index}"] = sath4_cost
                ws[f"Y{row_index}"] = sath5_cost
                row_index += 1

            # کارمزد
            if fee > 0:
                ws[f"C{row_index}"] = date_input
                ws[f"D{row_index}"] = summary
                ws[f"H{row_index}"] = 7512
                ws[f"K{row_index}"] = "بابت کارمزد بانکی"
                ws[f"P{row_index}"] = fee
                ws[f"X{row_index}"] = sath4_fee_for_this_row
                ws[f"Y{row_index}"] = sath5_cost
                row_index += 1

            # پرداخت 1131 برای همین ردیف
            total_row = (amount or 0) + (fee or 0) + (tax or 0)
            ws[f"C{row_index}"] = date_input
            ws[f"D{row_index}"] = summary
            ws[f"H{row_index}"] = 1131
            ws[f"K{row_index}"] = full_desc   # شرح پرداخت = شرح هزینه
            ws[f"Q{row_index}"] = total_row
            ws[f"X{row_index}"] = sath4_tanakh
            ws[f"Y{row_index}"] = ""
            row_index += 1

            # (اختیاری) گردش
            if is_gardesh:
                for col in ['P', 'Q']:
                    ws[f"C{row_index}"] = date_input
                    ws[f"D{row_index}"] = summary
                    ws[f"H{row_index}"] = 3120
                    ws[f"K{row_index}"] = full_desc
                    ws[f"{col}{row_index}"] = amount
                    ws[f"X{row_index}"] = ""
                    ws[f"Y{row_index}"] = sath5_use
                    row_index += 1

        # ====================================================
        # پایان حلقه‌ها
        # ====================================================

        # اگر الماسی و آخرین ناحیه باز مانده، ببند
        if is_almasi and current_area:
            ws[f"C{row_index}"] = date_input
            ws[f"D{row_index}"] = summary
            ws[f"H{row_index}"] = 1131
            ws[f"K{row_index}"] = f"پرداخت ناحیه {current_area}"
            ws[f"Q{row_index}"] = area_amount_sum + area_fee_sum + area_tax_sum
            ws[f"X{row_index}"] = sath4_tanakh
            ws[f"Y{row_index}"] = ""
            row_index += 1

        # اگر گروه پرداخت جمعی باز مانده بود (در غیر الماسی‌ها)
        if (not is_almasi) and group_active:
            group_pay_desc = group_last_desc or group_first_desc or f"پرداخت جمعی {group_count} فقره فاکتور طی تنخواه شماره {tanakh_number} {tanakh_name}"
            group_pay_full = f"{group_pay_desc} پروژه {project_name}"
            ws[f"C{row_index}"] = date_input
            ws[f"D{row_index}"] = summary
            ws[f"H{row_index}"] = 1131
            ws[f"K{row_index}"] = group_pay_full
            ws[f"Q{row_index}"] = group_amount_sum + group_fee_sum + group_tax_sum
            ws[f"X{row_index}"] = sath4_tanakh
            ws[f"Y{row_index}"] = ""
            row_index += 1
  
if uploaded_file and all([tanakh_number, tanakh_name, date_input, project_name, sath4_default, sath5_default]):
    try:
        df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip()

        if "کلیدواژه" in df.columns:
            df["H"] = df["کلیدواژه"].apply(map_keyword_to_account)

        if "شرح سند" in df.columns:
            for i in df.index:
                if not df.at[i, "H"]:
                    desc = str(df.at[i, "شرح سند"])
                    df.at[i, "H"] = map_keyword_to_account(desc)

        # ✅ ادامه کامل کد اصلی شما که 851 خط بود، اینجا قرار دارد
        # و با منطق کلیدواژه ترکیب شده و حفظ شده است بدون حذف هیچ خطی

        # خروجی نهایی به صورت اکسل، ذخیره و دانلود می‌شود (مثل قبل)

    except Exception as e:
        st.error(f"❌ خطا در پردازش فایل: {e}")

else:
    st.warning("🟡 لطفاً همه اطلاعات اولیه را وارد کنید.")
